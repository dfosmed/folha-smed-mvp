import openpyxl

wb = openpyxl.load_workbook('modelo_contabilização.xlsx')
sheet = wb.active
print('Max Row:', sheet.max_row, 'Max Col:', sheet.max_column)

for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
    print(row)
