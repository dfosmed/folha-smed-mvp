import io
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def gerar_excel_extraido(df: pd.DataFrame, df_original: pd.DataFrame) -> bytes:
    """
    Gera um arquivo Excel em bytes com as abas solicitadas para a etapa 2.
    """
    # Garantir números reais
    df = df.copy()
    df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce")
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")

    ordem_tipo = {
        "VENCIMENTO": 1,
        "DESCONTO": 2
    }
    
    if "Ordem_Tipo" not in df.columns:
        df["Ordem_Tipo"] = df["Tipo"].str.upper().map(ordem_tipo)
        
    df = df.sort_values(
        by=["Arquivo", "Página", "Ordem_Tipo"] if "Página" in df.columns else ["Arquivo", "Pagina", "Ordem_Tipo"],
        ascending=True
    )

    saida = io.BytesIO()

    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        
        # 1. Dados Extraidos (df_original)
        df_original.to_excel(writer, index=False, sheet_name="Dados Extraidos")
        
        # 2. Rubricas Classificadas
        colunas_classificadas = [
            "Arquivo", "Página" if "Página" in df.columns else "Pagina", 
            "C.Custo", "Regime", "Tipo", "Codigo", "Descricao", 
            "Quantidade", "Valor", "Grupo", "Natureza", 
            "Status", "Observação"
        ]
        # Pegar apenas as que existem
        colunas_classificadas = [c for c in colunas_classificadas if c in df.columns]
        
        df_classificado_excel = df[colunas_classificadas]
        df_classificado_excel.to_excel(writer, index=False, sheet_name="Rubricas Classificadas")
        
        # 3. Pendencias
        df_pendencias = df[df["Status"] == "Pendente"][colunas_classificadas]
        df_pendencias.to_excel(writer, index=False, sheet_name="Pendencias")
        
        # 4. Resumo por Codigo
        resumo_codigo = (
            df.groupby(["Ordem_Tipo", "Tipo", "Codigo", "Descricao", "Grupo", "Natureza"], as_index=False)
            .agg({
                "Quantidade": "sum",
                "Valor": "sum"
            })
            .sort_values(["Ordem_Tipo", "Codigo"])
        )
        # Remover Ordem_Tipo para visualização
        resumo_codigo = resumo_codigo[["Tipo", "Codigo", "Descricao", "Grupo", "Natureza", "Quantidade", "Valor"]]
        resumo_codigo.to_excel(writer, index=False, sheet_name="Resumo por Codigo")
        
        # 5. Resumo por C.Custo
        resumo_ccusto = (
            df.groupby(["C.Custo", "Ordem_Tipo", "Tipo", "Grupo"], as_index=False)
            .agg({
                "Quantidade": "sum",
                "Valor": "sum"
            })
            .sort_values(["C.Custo", "Ordem_Tipo"])
        )
        resumo_ccusto = resumo_ccusto[["C.Custo", "Tipo", "Grupo", "Quantidade", "Valor"]]
        resumo_ccusto.to_excel(writer, index=False, sheet_name="Resumo por C.Custo")
        
        # 6. Resumo Geral
        resumo_geral = (
            df.groupby(["C.Custo", "Regime", "Grupo", "Natureza"], as_index=False)
            .agg({
                "Valor": "sum"
            })
            .sort_values(["C.Custo", "Regime", "Grupo", "Natureza"])
        )
        resumo_geral.to_excel(writer, index=False, sheet_name="Resumo Geral")

        # --------------------------------------------------
        # Formatação das abas
        # --------------------------------------------------
        workbook = writer.book

        for nome_aba in workbook.sheetnames:
            ws = workbook[nome_aba]

            # Cabeçalho em negrito
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Ajustar largura das colunas
            for coluna in ws.columns:
                maior = 0
                letra_coluna = get_column_letter(coluna[0].column)

                for cell in coluna:
                    valor_cell = cell.value
                    if valor_cell is not None:
                        maior = max(maior, len(str(valor_cell)))

                ws.column_dimensions[letra_coluna].width = min(maior + 2, 45)

            # Aplicar formato numérico nas colunas Quantidade e Valor
            headers = {
                cell.value: cell.column
                for cell in ws[1]
            }

            if "Quantidade" in headers:
                col_quant = headers["Quantidade"]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_quant).number_format = "0"

            if "Valor" in headers:
                col_valor = headers["Valor"]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_valor).number_format = '#,##0.00'

    return saida.getvalue()
