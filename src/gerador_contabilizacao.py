import io
import re
import os
import json
import unicodedata
import pandas as pd
import openpyxl
import logging

logger = logging.getLogger(__name__)

CONFIG_BLOCOS = {
    "OUTROS SMED - EFETIVOS - CREDOR 105": {
        "C.Custo": "038 - OUTROS SMED",
        "Regimes": ["002 - EFETIVO", "015 - EFETIVO/COMISSIONADO", "019 - PROFISSIONAL EDUCACAO ( EST.)", "031 - PROF EDUCACAO EST./COMISSIONADO (CARREIRA)"],
        "Dotacao_CCusto": "OUTROS SMED",
        "Dotacao_Regime": "EFETIVO"
    },
    "OUTROS SMED - CONTRATADOS - CREDOR 543": {
        "C.Custo": "038 - OUTROS SMED",
        "Regimes": ["009 - CONTRATO", "020 - PROFISSIONAL EDUCACAO (CONT.)"],
        "Dotacao_CCusto": "OUTROS SMED",
        "Dotacao_Regime": "CONTRATADO"
    },
    "OUTROS SMED - COMISSIONADOS - CREDOR 543": {
        "C.Custo": "038 - OUTROS SMED",
        "Regimes": ["003 - COMISSIONADO"],
        "Dotacao_CCusto": "OUTROS SMED",
        "Dotacao_Regime": "COMISSIONADOS"
    },
    "OUTROS SMED - AGENTE POLÍTICO - CREDOR 543": {
        "C.Custo": "038 - OUTROS SMED",
        "Regimes": ["011 - AGENTE POLITICO"],
        "Dotacao_CCusto": "OUTROS SMED",
        "Dotacao_Regime": "AGENTE POLÍTICO"
    },
    "RESTANTE DA SMED 30% - COMISSIONADOS/CONTR. - CREDOR 543": {
        "C.Custo": "098 - RESTANTE SMED 30%",
        "Regimes": ["003 - COMISSIONADO"],
        "Dotacao_CCusto": "RESTANTE DA SMED 30%",
        "Dotacao_Regime": "COMISSIONADOS"
    },
    "RESTANTE DA SMED 30% - EFETIVOS - CREDOR 105": {
        "C.Custo": "098 - RESTANTE SMED 30%",
        "Regimes": ["015 - EFETIVO/COMISSIONADO", "019 - PROFISSIONAL EDUCACAO ( EST.)"],
        "Dotacao_CCusto": "RESTANTE DA SMED 30%",
        "Dotacao_Regime": "EFETIVO"
    },
    "RESTANTE DA SMED 30% - CONTRATADOS - CREDOR 543": {
        "C.Custo": "098 - RESTANTE SMED 30%",
        "Regimes": ["009 - CONTRATO", "020 - PROFISSIONAL EDUCACAO (CONT.)"],
        "Dotacao_CCusto": "RESTANTE DA SMED 30%",
        "Dotacao_Regime": "CONTRATADO"
    },
    "ED.FUNDAMENTAL 70% - COMISSIONADO- CREDOR 543": {
        "C.Custo": "093 - ENSINO FUNDAMENTAL 70%",
        "Regimes": ["003 - COMISSIONADO"],
        "Dotacao_CCusto": "ENS. FUNDAM. 70%",
        "Dotacao_Regime": "COMISSIONADOS"
    },
    "ED.FUNDAMENTAL 70%  - EFETIVOS - CREDOR 105": {
        "C.Custo": "093 - ENSINO FUNDAMENTAL 70%",
        "Regimes": ["002 - EFETIVO", "015 - EFETIVO/COMISSIONADO", "019 - PROFISSIONAL EDUCACAO ( EST.)", "030 - PROF EDUCACAO EST./COMISSIONADO", "031 - PROF EDUCACAO EST./COMISSIONADO (CARREIRA)", "032 - EFETIVO/COMISSIONADO (CARREIRA)"],
        "Dotacao_CCusto": "ENS. FUNDAM. 70%",
        "Dotacao_Regime": "EFETIVO"
    },
    "ED.FUNDAMENTAL 70% - CONTRATADOS - CREDOR 543": {
        "C.Custo": "093 - ENSINO FUNDAMENTAL 70%",
        "Regimes": ["009 - CONTRATO", "020 - PROFISSIONAL EDUCACAO (CONT.)"],
        "Dotacao_CCusto": "ENS. FUNDAM. 70%",
        "Dotacao_Regime": "CONTRATADO"
    },
    "ENSINO FUNDAMENTAL  30% EFETIVOS - CREDOR 105": {
        "C.Custo": "092 - ENSINO FUNDAMENTAL 30%",
        "Regimes": ["002 - EFETIVO"],
        "Dotacao_CCusto": "ENS. FUNDAM. 30%",
        "Dotacao_Regime": "EFETIVO"
    },
    "ENSINO FUNDAMENTAL 30% - CONTRATADOS - CREDOR 543": {
        "C.Custo": "092 - ENSINO FUNDAMENTAL 30%",
        "Regimes": ["009 - CONTRATO"],
        "Dotacao_CCusto": "ENS. FUNDAM. 30%",
        "Dotacao_Regime": "CONTRATADO"
    },
    "ED.INFANTIL  PRE ESCOLA 70% - EFETIVOS - CREDOR 105": {
        "C.Custo": "083 - E.I. PRÉ-ESCOLA 70%",
        "Regimes": ["002 - EFETIVO", "015 - EFETIVO/COMISSIONADO", "019 - PROFISSIONAL EDUCACAO ( EST.)", "030 - PROF EDUCACAO EST./COMISSIONADO", "031 - PROF EDUCACAO EST./COMISSIONADO (CARREIRA)"],
        "Dotacao_CCusto": "E.I. PRE ESCOLA 70%",
        "Dotacao_Regime": "EFETIVO"
    },
    "ED.INFANTIL  PRE ESCOLA 70% - CONTRATADOS - CREDOR 543": {
        "C.Custo": "083 - E.I. PRÉ-ESCOLA 70%",
        "Regimes": ["009 - CONTRATO", "020 - PROFISSIONAL EDUCACAO (CONT.)"],
        "Dotacao_CCusto": "E.I. PRE ESCOLA 70%",
        "Dotacao_Regime": "CONTRATADO"
    },
    "ED.INFANTIL  PRE ESCOLA 30% EFETIVOS - CREDOR 105": {
        "C.Custo": "082 - E.I. PRÉ-ESCOLA 30%",
        "Regimes": ["002 - EFETIVO", "019 - PROFISSIONAL EDUCACAO ( EST.)"],
        "Dotacao_CCusto": "E.I. PRE ESCOLA 30%",
        "Dotacao_Regime": "EFETIVO"
    },
    "ED.INFANTIL CRECHE 70% - EFETIVOS - CREDOR 105": {
        "C.Custo": "077 - E.I. CRECHE 70%",
        "Regimes": ["002 - EFETIVO", "015 - EFETIVO/COMISSIONADO", "019 - PROFISSIONAL EDUCACAO ( EST.)", "030 - PROF EDUCACAO EST./COMISSIONADO", "031 - PROF EDUCACAO EST./COMISSIONADO (CARREIRA)", "032 - EFETIVO/COMISSIONADO (CARREIRA)"],
        "Dotacao_CCusto": "E.I. CRECHE 70%",
        "Dotacao_Regime": "EFETIVO"
    },
    "ED.INFANTIL  CRECHE 70% - CONTRATADOS - CREDOR 543": {
        "C.Custo": "077 - E.I. CRECHE 70%",
        "Regimes": ["009 - CONTRATO", "020 - PROFISSIONAL EDUCACAO (CONT.)"],
        "Dotacao_CCusto": "E.I. CRECHE 70%",
        "Dotacao_Regime": "CONTRATADO"
    },
    "ED.INFANTIL  CRECHE 30% EFETIVOS - CREDOR 105": {
        "C.Custo": "076 - E.I. CRECHE 30%",
        "Regimes": ["002 - EFETIVO", "019 - PROFISSIONAL EDUCACAO ( EST.)", "030 - PROF EDUCACAO EST./COMISSIONADO"],
        "Dotacao_CCusto": "E.I. CRECHE 30%",
        "Dotacao_Regime": "EFETIVO"
    },
    "EDUCAÇÃO ESPECIAL 70% EFETIVOS - CREDOR 105": {
        "C.Custo": "072 - EDUCAÇÃO ESPECIAL 70%",
        "Regimes": ["002 - EFETIVO", "019 - PROFISSIONAL EDUCACAO ( EST.)"],
        "Dotacao_CCusto": "ED. ESPECIAL 70%",
        "Dotacao_Regime": "EFETIVO"
    },
    "EDUCAÇÃO ESPECIAL 70% - CONTRATADOS - CREDOR 543": {
        "C.Custo": "072 - EDUCAÇÃO ESPECIAL 70%",
        "Regimes": ["009 - CONTRATO", "020 - PROFISSIONAL EDUCACAO (CONT.)"],
        "Dotacao_CCusto": "ED. ESPECIAL 70%",
        "Dotacao_Regime": "CONTRATADO"
    },
    "EJA 70%  EFETIVOS - CREDOR 105": {
        "C.Custo": "088 - EJA 70%",
        "Regimes": ["002 - EFETIVO", "015 - EFETIVO/COMISSIONADO", "019 - PROFISSIONAL EDUCACAO ( EST.)", "031 - PROF EDUCACAO EST./COMISSIONADO (CARREIRA)"],
        "Dotacao_CCusto": "EJA 70%",
        "Dotacao_Regime": "EFETIVO"
    },
    "EJA 70% - CONTRATADOS - CREDOR 543": {
        "C.Custo": "088 - EJA 70%",
        "Regimes": ["009 - CONTRATO", "020 - PROFISSIONAL EDUCACAO (CONT.)"],
        "Dotacao_CCusto": "EJA 70%",
        "Dotacao_Regime": "CONTRATADO"
    }
}

NATUREZA_MAP = {
    'ABONO FAMILIA': 'ABONO FAMÍLIA',
    'SALÁRIO FAMILIA (INSS/CONTR.)': 'SALÁRIO FAMÍLIA',
    'AFAST.MATERNIDADE (SAL. MATERN.)': 'AFASTAMENTO MATERNIDADE',
    'AFAST. MATERNIDADE(INSS/CONTR.)': 'AFASTAMENTO MATERNIDADE',
    'AUXILIO DOENÇA (LICENÇA SAÚDE)': 'AUXÍLIO DOENÇA',
    'FÉRIAS 1/3 (ABONO CONSTITUCIONAL)': 'FÉRIAS - ABONO CONSTITUCIONAL',
    'FÉRIAS 1/3  (ABONO CONSTITUCIONAL)': 'FÉRIAS - ABONO CONSTITUCIONAL',
    'PARCELA PROP. (13ºSLR)': '13º SALÁRIO',
    'HORA EXTRA': 'HORA EXTRA',
    'INDENIZ. E RESTITUIÇÕES TRAB.': 'IDENIZAÇÕES E RESTITUIÇÕES',
    'AUXÍLIO NATALIDADE': 'AUXÍLIO NATALIDADE',
    'RESSARCIMENTO': 'RESSARCIMENTO',
    'DEDUÇÃO ART. 37, X': 'DEDUÇÃO ART. 37, X',
    'DIFERENÇA CARGA HORÁRIA': 'DIFERENÇA CARGA HORÁRIA',
    'DESCONTO DIAS/HORAS': 'DESCONTO DIAS/HORAS',
    'FALTAS/FALTAS HORAS': 'FALTAS/FALTAS HORAS',
}

def normalize_text(text):
    if not isinstance(text, str):
        return ""
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')
    text = text.upper()
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def normalize_key(text):
    text = normalize_text(text)
    text = re.sub(r'[^A-Z0-9]', '', text)
    return text

def get_dotacao_value(dotacoes_json, ccusto, regime, natureza_planilha, tipo_dotacao):
    if dotacoes_json is None or not tipo_dotacao:
        return None
        
    natureza_norm = normalize_key(natureza_planilha)
    tipo_norm = normalize_key(tipo_dotacao)
    
    # Encontrar no json iterando
    for row_data in dotacoes_json:
        r_ccusto = str(row_data.get('C. CUSTO', '')).strip().upper()
        r_regime = str(row_data.get('REGIME', '')).strip().upper()
        r_desc = str(row_data.get('DESCRIÇÃO', row_data.get('DESCRIÇO', ''))).strip() 
        
        if r_ccusto == ccusto.strip().upper() and r_regime == regime.strip().upper():
            if normalize_key(r_desc) == natureza_norm:
                # Encontrou a linha, agora acha a coluna
                for col, val in row_data.items():
                    if normalize_key(col) == tipo_norm:
                        if val is not None and not (isinstance(val, float) and pd.isna(val)):
                            if str(val).strip().lower() != "nan":
                                return str(val).strip()
                break
    return None

def parse_and_fill_contabilizacao(df_resumo: pd.DataFrame, path_modelo, config_dotacoes_dict=None) -> bytes:
    if config_dotacoes_dict is None:
        config_dotacoes_dict = {}
    if isinstance(path_modelo, str):
        wb = openpyxl.load_workbook(path_modelo)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(path_modelo))
    sheet = wb.active
    
    # Carregar detalhamento de dotações via JSON
    dotacoes_json = None
    try:
        json_path = os.path.join(os.path.dirname(__file__), 'dotacoes.json')
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                dotacoes_json = json.load(f)
    except Exception as e:
        logger.error(f"Não foi possivel carregar Detalhamento de Dotações JSON: {e}")
    
    # Precompute keys for block identification
    block_keys = {normalize_key(k): v for k, v in CONFIG_BLOCOS.items()}
    
    # Dicionario para acesso rápido às naturezas pelo texto
    natureza_map_norm = {normalize_key(k): v for k, v in NATUREZA_MAP.items()}
    
    current_block_config = None
    current_block_df = None
    start_data_row = None
    
    in_block = False
    folha_bruta_row = None
    mapped_naturezas_in_block = set()
    
    for row in range(1, sheet.max_row + 1):
        cell_val = sheet.cell(row=row, column=1).value
        if not cell_val:
            continue
            
        norm_cell = normalize_key(str(cell_val))
        
        # Detect block start
        if norm_cell in block_keys:
            current_block_config = block_keys[norm_cell]
            # Extrair df
            c_custo = current_block_config["C.Custo"]
            regimes = current_block_config["Regimes"]
            current_block_df = df_resumo[(df_resumo['C.Custo'] == c_custo) & (df_resumo['Regime'].isin(regimes))]
            
            in_block = True
            folha_bruta_row = None
            start_data_row = None
            mapped_naturezas_in_block = set()
            continue
            
        if in_block:
            dotacao_atual = sheet.cell(row=row, column=3).value
            manter_dotacao = False
            
            # Força o NÃO EMPENHAR para rubricas específicas que sempre o utilizam
            rubricas_nao_empenhar = [
                "PARCELAANT13", "RESSARCIMENTO", "DEDUCAOART37X", 
                "DIFERENCACARGAHORARIA", "DESCONTODIASHORAS", "FALTASFALTASHORAS"
            ]
            if norm_cell in rubricas_nao_empenhar:
                # Trata erro de merged cell
                if type(sheet.cell(row=row, column=3)).__name__ != 'MergedCell':
                    sheet.cell(row=row, column=3).value = "NÃO EMPENHAR"
                    import copy
                    if sheet.cell(row=row, column=3).font:
                        new_font = copy.copy(sheet.cell(row=row, column=3).font)
                        new_font.bold = True
                        sheet.cell(row=row, column=3).font = new_font
                manter_dotacao = True
                
            if dotacao_atual and ("NǟO EMPENHAR" in str(dotacao_atual).upper() or "NAO EMPENHAR" in str(dotacao_atual).upper() or "NÃO EMPENHAR" in str(dotacao_atual).upper()):
                manter_dotacao = True
                
            # Look for Folha Bruta
            if "FOLHABRUTA" in norm_cell or "BRUTOSUBISDIO" in norm_cell or "BRUTOSUBSIDIO" in norm_cell or ("BRUTO" in norm_cell and "SECRETARIO" in norm_cell):
                if not folha_bruta_row and "EMPENHO" not in norm_cell:
                    folha_bruta_row = row
                    
                    if current_block_config.get("Dotacao_Regime") == "AGENTE POLÍTICO":
                        val_folha = current_block_df[current_block_df['Natureza'] == 'VENCIMENTOS']['Valor'].sum()
                    else:
                        val_folha = current_block_df[current_block_df['Grupo'] == 'Provento']['Valor'].sum()
                        
                    if val_folha > 0:
                        sheet.cell(row=row, column=2).value = val_folha
                    else:
                        sheet.cell(row=row, column=2).value = None 
                    continue
            
            # Look for block end / Folha Bruta para Empenho
            if "EMPENHO" in norm_cell and (("FOLHA" in norm_cell and "BRUTA" in norm_cell) or "SUBSIDIO" in norm_cell):
                if current_block_config.get("Dotacao_Regime") == "AGENTE POLÍTICO":
                    val_folha_empenho = current_block_df[current_block_df['Natureza'] == 'VENCIMENTOS']['Valor'].sum()
                else:
                    val_folha_empenho = current_block_df[current_block_df['Grupo'] == 'Provento']['Valor'].sum()
                    
                unmapped_df = current_block_df[(current_block_df['Grupo'] == 'Desconto') & (~current_block_df['Natureza'].isin(mapped_naturezas_in_block))]
                sum_unmapped = unmapped_df['Valor'].sum()
                
                if val_folha_empenho > 0 and sum_unmapped > 0:
                    sheet.cell(row=row, column=4).value = sum_unmapped
                else:
                    sheet.cell(row=row, column=4).value = None
                    
                nova_dotacao_empenho = None
                if dotacoes_json is not None:
                    d_ccusto = current_block_config.get("Dotacao_CCusto")
                    d_regime = current_block_config.get("Dotacao_Regime")
                    if d_ccusto and d_regime:
                        key = f"{d_ccusto.upper()}_{d_regime.upper()}"
                        tipo_selecionado = config_dotacoes_dict.get(key, None)
                        # Busca do JSON a rubrica "FOLHA BRUTA PARA EMPENHO"
                        nova_dotacao_empenho = get_dotacao_value(dotacoes_json, d_ccusto, d_regime, "FOLHA BRUTA PARA EMPENHO", tipo_selecionado)

                is_special_dynamic_empenho = nova_dotacao_empenho and nova_dotacao_empenho.upper() in ["OP EXTRA", "NÃO EMPENHAR", "NAO EMPENHAR"]
                
                def set_dotacao_empenho(val, bold=False):
                    if type(sheet.cell(row=row, column=3)).__name__ != 'MergedCell':
                        sheet.cell(row=row, column=3).value = val
                        import copy
                        if sheet.cell(row=row, column=3).font:
                            new_font = copy.copy(sheet.cell(row=row, column=3).font)
                            new_font.bold = bold
                            sheet.cell(row=row, column=3).font = new_font
                
                if val_folha_empenho > 0 and sum_unmapped > 0:
                    if is_special_dynamic_empenho:
                        set_dotacao_empenho(nova_dotacao_empenho, bold=True)
                    elif nova_dotacao_empenho:
                        set_dotacao_empenho(nova_dotacao_empenho, bold=False)
                    else:
                        set_dotacao_empenho(None, bold=False)
                else:
                    if is_special_dynamic_empenho:
                        set_dotacao_empenho(nova_dotacao_empenho, bold=True)
                    else:
                        set_dotacao_empenho(None, bold=False)
                
                in_block = False
                continue
                
            # Else, see if it's a known mapped row
            if norm_cell in natureza_map_norm:
                expected_nat = natureza_map_norm[norm_cell]
                val_nat = current_block_df[current_block_df['Natureza'] == expected_nat]['Valor'].sum()
                
                nova_dotacao = None
                if dotacoes_json is not None:
                    d_ccusto = current_block_config.get("Dotacao_CCusto")
                    d_regime = current_block_config.get("Dotacao_Regime")
                    if d_ccusto and d_regime:
                        key = f"{d_ccusto.upper()}_{d_regime.upper()}"
                        tipo_selecionado = config_dotacoes_dict.get(key, None)
                        nova_dotacao = get_dotacao_value(dotacoes_json, d_ccusto, d_regime, cell_val, tipo_selecionado)
                        
                is_special_dynamic = nova_dotacao and nova_dotacao.upper() in ["OP EXTRA"]
                is_special_rubrica = norm_cell in rubricas_nao_empenhar
                
                def set_dotacao(val, bold=False):
                    if type(sheet.cell(row=row, column=3)).__name__ != 'MergedCell':
                        sheet.cell(row=row, column=3).value = val
                        import copy
                        if sheet.cell(row=row, column=3).font:
                            new_font = copy.copy(sheet.cell(row=row, column=3).font)
                            new_font.bold = bold
                            sheet.cell(row=row, column=3).font = new_font
                            
                if val_nat > 0:
                    sheet.cell(row=row, column=2).value = val_nat
                    mapped_naturezas_in_block.add(expected_nat)
                    
                    if is_special_rubrica:
                        pass # Já forçado no início do bloco
                    elif is_special_dynamic:
                        set_dotacao(nova_dotacao, bold=True)
                    elif nova_dotacao:
                        set_dotacao(nova_dotacao, bold=False)
                    else:
                        set_dotacao(None, bold=False)
                else:
                    sheet.cell(row=row, column=2).value = None
                    if is_special_rubrica:
                        pass # Já forçado no início do bloco
                    elif is_special_dynamic:
                        set_dotacao(nova_dotacao, bold=True)
                    else:
                        set_dotacao(None, bold=False)
                            
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
